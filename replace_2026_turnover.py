import openpyxl
from datetime import datetime, timedelta

TEMPLATE = r'C:\Users\zhangyufei1\Desktop\更新部署\7月库存分析模版.xlsx'
INV = r'C:\Users\zhangyufei1\WorkBuddy\2026-06-30-09-24-40\inventory.xlsx'

RDC_SHORT = ['东北', '华北', '华东', '华南', '华中', '西北', '西南']

def s2date(v):
    if isinstance(v, (int, float)) and v > 40000:
        return (datetime(1899, 12, 30) + timedelta(days=v)).strftime('%Y-%m')
    if isinstance(v, datetime):
        return v.strftime('%Y-%m')
    return str(v)

# ---------- read template ----------
twb = openpyxl.load_workbook(TEMPLATE, read_only=True, data_only=True)
src = twb['2026周转']
trows = [list(r) for r in src.iter_rows(values_only=True)]

# month serials (left sub-header row 1, cols 1-7)
left_serials = [trows[1][c] for c in range(1, 8)]
right_serials = [trows[1][c] for c in range(10, 17)]
print('left months :', [s2date(x) for x in left_serials])
print('right months:', [s2date(x) for x in right_serials])

# split marker (出货成本统计) row
splitIdx = next((i for i, r in enumerate(trows) if r and r[0] and '出货成本统计' in str(r[0])), len(trows))
print('出货成本统计 row index:', splitIdx)

# sub-location marker (RDC共享库位库存周转天数)
subMarker = next((i for i, r in enumerate(trows) if r and r[0] and 'RDC共享库位库存周转天数' in str(r[0])), len(trows))
print('共享库位 marker row index:', subMarker)

def g(r, c):
    return r[c] if r and len(r) > c else None

# left turnover by label (cols 1-7, 7 months)
allLoc = {}        # '东北RDC' -> [7]
totalAll = None
totalShared = None
hq = None
subLoc = {}        # '东北RDC' -> [7]
for i, r in enumerate(trows):
    lbl = str(r[0]).strip() if r and r[0] else ''
    vals7 = [g(r, c) for c in range(1, 8)]
    if lbl == 'RDC（全库位）': totalAll = vals7
    elif lbl == 'RDC（共享仓）': totalShared = vals7
    elif lbl == '总仓': hq = vals7
    elif lbl.endswith('RDC分仓'):
        name = lbl.replace('RDC分仓', 'RDC')
        allLoc[name] = vals7
    elif i > subMarker and lbl in RDC_SHORT:
        subLoc[lbl + 'RDC'] = vals7

# right inventory (pre-split) and shipment (post-split) by RDC short name (cols 10-16)
invByRdc = {}
shipByRdc = {}
for i, r in enumerate(trows):
    lbl = str(g(r, 9)).strip() if g(r, 9) else ''
    if lbl in RDC_SHORT:
        vals7 = [g(r, c) for c in range(10, 17)]
        if i < splitIdx:
            invByRdc[lbl] = vals7
        else:
            shipByRdc[lbl] = vals7

print('inventory RDCs:', sorted(invByRdc.keys()))
print('shipment RDCs:', sorted(shipByRdc.keys()))
print('allLoc RDCs  :', sorted(allLoc.keys()))
print('subLoc RDCs  :', sorted(subLoc.keys()))
# sanity: 东北 July inventory blank?
print('东北 inv July:', invByRdc.get('东北', [None]*7)[6], '| 华北 inv July:', invByRdc.get('华北', [None]*7)[6])
print('华东 allLoc July:', allLoc.get('华东RDC', [None]*7)[6], '(华东停运应为空)')

# ---------- build output rows matching CURRENT clean structure (7 months) ----------
def row(*cells):
    out = [None] * 16
    for i, v in enumerate(cells):
        out[i] = v
    return out

M = left_serials  # left month serials
MR = right_serials # right month serials

out = []
out.append(row('RDC周转天数总览', '', '', '', '', '', '', '', '库存金额统计（单位：千元）', '', '', '', '', '', '', '', '', ''))
out.append(row('仓库', *M, '', '仓库', *MR))  # r1
out.append(row('RDC（全库位）', *totalAll, '', '东北', *invByRdc['东北']))            # r2
out.append(row('RDC（共享仓）', *totalShared, '', '华北', *invByRdc['华北']))          # r3
out.append(row('总仓', *hq, '', '华东', *invByRdc['华东']))                            # r4
out.append(row('', '', '', '', '', '', '', '', '华南', *invByRdc['华南']))            # r5
out.append(row('', '', '', '', '', '', '', '', '华中', *invByRdc['华中']))            # r6
out.append(row('RDC全库位库存周转天数', '', '', '', '', '', '', '', '西北', *invByRdc['西北']))  # r7
out.append(row('仓库', *M, '', '西南', *invByRdc['西南']))                            # r8
out.append(row('东北RDC分仓', *allLoc['东北RDC'], '', '总仓', *invByRdc['总仓']))      # r9
out.append(row('华北RDC分仓', *allLoc['华北RDC'], '', '', ''))                        # r10
out.append(row('华南RDC分仓', *allLoc['华南RDC'], '', '', ''))                        # r11
out.append(row('华中RDC分仓', *allLoc['华中RDC'], '', '出货成本统计（单位：千元）', ''))  # r12 splitPoint
out.append(row('西北RDC分仓', *allLoc['西北RDC'], '', '仓库', *MR))                   # r13
out.append(row('西南RDC分仓', *allLoc['西南RDC'], '', '东北', *shipByRdc['东北']))     # r14
out.append(row('华东RDC分仓', *allLoc['华东RDC'], '', '华北', *shipByRdc['华北']))     # r15
out.append(row('', '', '', '', '', '', '', '', '华东', *shipByRdc['华东']))           # r16
out.append(row('', '', '', '', '', '', '', '', '华南', *shipByRdc['华南']))           # r17
out.append(row('RDC共享库位库存周转天数', '', '', '', '', '', '', '', '华中', *shipByRdc['华中']))  # r18
out.append(row('仓库', *M, '', '西北', *shipByRdc['西北']))                           # r19
out.append(row('东北', *subLoc['东北RDC'], '', '西南', *shipByRdc['西南']))           # r20
out.append(row('华北', *subLoc['华北RDC'], '', '总仓', *shipByRdc['总仓']))           # r21
out.append(row('华南', *subLoc['华南RDC'], '', '', ''))                               # r22
out.append(row('华中', *subLoc['华中RDC'], '', '', ''))                               # r23
out.append(row('西北', *subLoc['西北RDC'], '', '', ''))                               # r24
out.append(row('西南', *subLoc['西南RDC'], '', '', ''))                               # r25
out.append(row('华东', *subLoc['华东RDC'], '', '周转天数=', '月末存货金额 / YTD销售成本 * YTD天数(每月按30天计)'))  # r26
for _ in range(27, 34):
    out.append(row())

# ---------- replace in inventory.xlsx (preserve order) ----------
iwb = openpyxl.load_workbook(INV)
if '2026周转' in iwb.sheetnames:
    del iwb['2026周转']
ws = iwb.create_sheet('2026周转')
for r in out:
    ws.append(r)
order = ['订单满足率', '海南花露水历史数据', '2026周转', '2025周转', '总体情况',
         '5月库存状态分析', '6月库存状态分析', '5月库存覆盖数据', '6月库存覆盖数据',
         '拉回数据', '转储数据', '基础数据']
iwb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
iwb.save(INV)
print('saved', INV)

# ---------- verify ----------
vw = openpyxl.load_workbook(INV, read_only=True, data_only=True)
vr = [list(r) for r in vw['2026周转'].iter_rows(values_only=True)]
print('verify rows:', len(vr))
print('r1 right months:', [s2date(x) for x in vr[1][9:16]])
print('r2 东北 inv Jul(col15):', vr[2][15], '| r3 华北 inv Jul:', vr[3][15], '| r9 总仓 inv Jul:', vr[9][15])
print('r14 东北 ship Jul:', vr[14][15], '| r15 华北 ship Jul:', vr[15][15])
print('r12 col8 (split):', vr[12][8])
print('r2 left Jul(col7) RDC全库位:', vr[2][7])
