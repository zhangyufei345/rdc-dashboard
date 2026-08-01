#!/usr/bin/env python3
# 补全 拉回数据 7月行的 品牌(col8)：源模版品牌列为空，用 基础数据 的 物料->品牌 映射回填，
# 保持与既有(1-6月)拉回数据的同口径，避免品牌维度7月全显“未知”。
import openpyxl
from datetime import datetime, timedelta

INV = 'inventory.xlsx'
def s2date(v):
    if isinstance(v, datetime): return v.strftime('%Y-%m')
    if isinstance(v, (int, float)) and v > 40000:
        return (datetime(1899, 12, 30) + timedelta(days=v)).strftime('%Y-%m')
    if isinstance(v, str):
        s = v.strip()
        if len(s) >= 7 and s[4] == '-': return s[:7]
    return ''

cwb = openpyxl.load_workbook(INV)
# 物料->品牌 (基础数据 col0=产品编码, col4=品牌)
bws = cwb['基础数据']
mat2brand = {}
for row in bws.iter_rows(min_row=2, values_only=True):
    if not row or row[0] is None: continue
    mat = str(row[0]).strip()
    brand = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
    if mat and brand:
        mat2brand[mat] = brand

pws = cwb['拉回数据']
updated = 0
for row in pws.iter_rows(min_row=2):
    if not row or row[0].value is None: continue
    if s2date(row[0].value) != '2026-07': continue
    # col3=物料, col8=品牌
    mat = str(row[3].value).strip() if row[3].value is not None else ''
    if mat and (row[8].value is None or str(row[8].value).strip() == ''):
        b = mat2brand.get(mat)
        if b:
            row[8].value = b
            updated += 1
print(f'品牌补全: {updated} 行 (映射表命中 {len(mat2brand)} 个物料)')
cwb.save(INV)
print('已保存', INV)
