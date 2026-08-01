#!/usr/bin/env python3
# 将「7月库存分析模版.xlsx」中的 7月 转储数据 / 拉回数据 合并进当前 inventory.xlsx
# 约束：只更新这两个 Sheet，其他库存数据(覆盖/周转/状态)保持原样。
# 关键：拉回数据新格式无直接 RDC，需用 基础数据 的 发货工厂->RDC 映射解析。
import openpyxl
from datetime import datetime, timedelta

TEMPLATE = r'C:\Users\zhangyufei1\Desktop\更新部署\7月库存分析模版.xlsx'
INV = 'inventory.xlsx'

def excel_date(v):
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)) and v > 40000:
        return (datetime(1899, 12, 30) + timedelta(days=v)).strftime('%Y-%m-%d')
    if isinstance(v, str):
        s = v.strip()
        if len(s) >= 10 and s[4] == '-' and s[7] == '-': return s[:10]
        if len(s) >= 7 and s[4] == '-': return s[:7]
    return ''

print('读取模版...')
twb = openpyxl.load_workbook(TEMPLATE, read_only=True, data_only=True)

# ---- 基础数据: 发货工厂->RDC, 收货库位->收货仓 ----
tws = twb['基础数据']; trows = list(tws.iter_rows(values_only=True)); thdr = trows[0]
def ci(name):
    for i, h in enumerate(thdr):
        if isinstance(h, str) and h.strip() == name: return i
    return -1
C_LOC, C_WH, C_FAC, C_RDC = ci('收货库位'), ci('收货仓'), ci('发货工厂'), ci('RDC')
fac2rdc, loc2wh = {}, {}
for r in trows[1:]:
    if not r: continue
    fac = str(r[C_FAC]).strip() if C_FAC >= 0 and r[C_FAC] is not None else ''
    rdc = str(r[C_RDC]).strip() if C_RDC >= 0 and r[C_RDC] is not None else ''
    loc = str(r[C_LOC]).strip() if C_LOC >= 0 and r[C_LOC] is not None else ''
    wh  = str(r[C_WH]).strip()  if C_WH  >= 0 and r[C_WH]  is not None else ''
    if fac and rdc: fac2rdc[fac] = rdc
    if loc: loc2wh[loc] = wh
print('发货工厂->RDC:', fac2rdc)

# ---- 转储数据(仅7月) ----
tws = twb['转储数据']; trows = list(tws.iter_rows(values_only=True))
trans_new = []
for r in trows[1:]:
    if not r or r[1] is None: continue
    ds = excel_date(r[1])
    if not ds.startswith('2026-07'): continue
    trans_new.append(r)

# ---- 拉回数据(仅7月) 转换到当前9列格式 ----
# 模版: [创建日期, 收货库位, 物料, 短文本, 发货单数量, OUn, 发货工厂, 品牌]
# 当前: [创建日期, 收货库位, 收货仓, 物料, 短文本, 发货单数量, 单位, 发货RDC, 品牌]
tws = twb['拉回数据']; trows = list(tws.iter_rows(values_only=True))
pull_new, bad = [], 0
for r in trows[1:]:
    if not r or r[0] is None: continue
    ds = excel_date(r[0])
    if not ds.startswith('2026-07'): continue
    loc  = str(r[1]).strip() if r[1] is not None else ''
    mat  = str(r[2]).strip() if r[2] is not None else ''
    name = str(r[3]).strip() if r[3] is not None else ''
    qty  = r[4]
    unit = str(r[5]).strip() if r[5] is not None else ''
    fac  = str(r[6]).strip() if r[6] is not None else ''
    brand= str(r[7]).strip() if r[7] is not None else ''
    rdc  = fac2rdc.get(fac, '')
    wh   = loc2wh.get(loc, '')
    if not rdc: bad += 1
    pull_new.append([r[0], loc, wh, mat, name, qty, unit, rdc, brand])
print(f'7月 转储: {len(trans_new)} 行 | 7月 拉回: {len(pull_new)} 行 (RDC未解析: {bad})')

# ---- 加载当前 inventory.xlsx (正常模式以便写入) ----
print('加载当前 inventory.xlsx ...')
cwb = openpyxl.load_workbook(INV)
cws_t = cwb['转储数据']
cws_p = cwb['拉回数据']

# 现有 转储 key = (日期, 物料(col5), 收货库位(col3))
exist_t = set()
for row in cws_t.iter_rows(min_row=2, values_only=True):
    if not row or row[1] is None: continue
    exist_t.add((excel_date(row[1]), str(row[5] or '').strip(), str(row[3] or '').strip()))
added_t = 0
for r in trans_new:
    key = (excel_date(r[1]), str(r[5] or '').strip(), str(r[3] or '').strip())
    if key in exist_t: continue
    exist_t.add(key)
    cws_t.append(list(r)); added_t += 1

# 现有 拉回 key = (日期, 收货库位(col1), 物料(col3), 发货单数量(col5), 发货RDC(col7))
exist_p = set()
for row in cws_p.iter_rows(min_row=2, values_only=True):
    if not row or row[0] is None: continue
    exist_p.add((excel_date(row[0]), str(row[1] or '').strip(), str(row[3] or '').strip(),
                 str(row[5] or '').strip(), str(row[7] or '').strip()))
added_p = 0
for nr in pull_new:
    key = (excel_date(nr[0]), str(nr[1] or '').strip(), str(nr[3] or '').strip(),
           str(nr[5] or '').strip(), str(nr[7] or '').strip())
    if key in exist_p: continue
    exist_p.add(key)
    cws_p.append(nr); added_p += 1

print(f'转储 新增: {added_t} | 拉回 新增: {added_p}')
cwb.save(INV)
print('已保存', INV)

# ---- 校验: 重新读取确认7月已写入 ----
vwb = openpyxl.load_workbook(INV, read_only=True, data_only=True)
def cnt_month(sheet, col):
    from collections import Counter
    c = Counter()
    for r in vwb[sheet].iter_rows(values_only=True):
        if r and r[col] is not None:
            d = excel_date(r[col])
            if d: c[d[:7]] += 1
    return dict(c)
print('校验 转储 月份分布:', cnt_month('转储数据', 1))
print('校验 拉回 月份分布:', cnt_month('拉回数据', 0))
